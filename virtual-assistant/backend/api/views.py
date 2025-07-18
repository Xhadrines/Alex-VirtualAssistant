import os
import logging
import requests
from bs4 import BeautifulSoup
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from django.conf import settings
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.views.generic import TemplateView
from django.urls import reverse_lazy
from .models import (
    Facultate,
    Specializare,
    Grupa,
    UserProfile,
    Adevetinta,
    ConversationHistory,
)
from .serializers import (
    FacultateSerializer,
    SpecializareSerializer,
    GrupaSerializer,
    UserSerializer,
    UserProfileSerializer,
    AdevetintaSerializer,
    ConversationHistorySerializer,
)
from .rag import RAG, remove_diacritics
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from django.template.loader import render_to_string
from weasyprint import HTML
import uuid
from django.http import JsonResponse
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
from django.views import View
from django.utils.decorators import method_decorator
from django.core.exceptions import ObjectDoesNotExist
import shutil
from dotenv import load_dotenv

# TODO: Pentru a folosi filtrul pentru specializare si grupa la final trebuie introdus .../?[numele variabilei respectiva]=[id-ul respectiv]
# TODO: Pentru a descarca pdf-urile de pe site trebuie sa trimiti o cerere json goala "{}"

# ----------------------------------------------------------------------------------------------------

logger = logging.getLogger(__name__)

# ----------------------------------------------------------------------------------------------------


class FacultateListView(generics.ListAPIView):
    queryset = Facultate.objects.all()
    serializer_class = FacultateSerializer


class FacultateCreateView(generics.CreateAPIView):
    queryset = Facultate.objects.all()
    serializer_class = FacultateSerializer


class FacultateUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Facultate.objects.all()
    serializer_class = FacultateSerializer


class FacultateDeleteView(generics.DestroyAPIView):
    queryset = Facultate.objects.all()
    serializer_class = FacultateSerializer


class FacultateFilteredView(APIView):
    def get(self, request, *args, **kwargs):
        nume_facultate = request.query_params.get("nume")
        if nume_facultate:
            facultati = Facultate.objects.filter(nume__icontains=nume_facultate)
            serializer = FacultateSerializer(facultati, many=True)
            return Response(serializer.data)
        return Response(
            {"error": "Selectarea facultatii este necesar"},
            status=status.HTTP_400_BAD_REQUEST,
        )


# ----------------------------------------------------------------------------------------------------


class SpecializareListView(generics.ListAPIView):
    queryset = Specializare.objects.all()
    serializer_class = SpecializareSerializer


class SpecializareCreateView(generics.CreateAPIView):
    queryset = Specializare.objects.all()
    serializer_class = SpecializareSerializer


class SpecializareUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Specializare.objects.all()
    serializer_class = SpecializareSerializer


class SpecializareDeleteView(generics.DestroyAPIView):
    queryset = Specializare.objects.all()
    serializer_class = SpecializareSerializer


class SpecializareFilteredView(APIView):
    def get(self, request, *args, **kwargs):
        facultate = request.query_params.get("facultate")
        if facultate is not None:
            specializari = Specializare.objects.filter(facultate=facultate)
            serializer = SpecializareSerializer(specializari, many=True)
            return Response(serializer.data)
        return Response(
            {"error": "Selectarea facultatii este necesar"},
            status=status.HTTP_400_BAD_REQUEST,
        )


# -----------------------------------------------------------------------------------------------------


class GrupaListView(generics.ListAPIView):
    queryset = Grupa.objects.all()
    serializer_class = GrupaSerializer


class GrupaCreateView(generics.CreateAPIView):
    queryset = Grupa.objects.all()
    serializer_class = GrupaSerializer


class GrupaUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Grupa.objects.all()
    serializer_class = GrupaSerializer


class GrupaDeleteView(generics.DestroyAPIView):
    queryset = Grupa.objects.all()
    serializer_class = GrupaSerializer


class GrupaFilteredView(APIView):
    def get(self, request, *args, **kwargs):
        specializare = request.query_params.get("specializare")
        if specializare is not None:
            grupe = Grupa.objects.filter(specializare=specializare)
            serializer = GrupaSerializer(grupe, many=True)
            return Response(serializer.data)
        return Response(
            {"error": "Selectarea specializare este necesar"},
            status=status.HTTP_400_BAD_REQUEST,
        )


# ----------------------------------------------------------------------------------------------------


class UserListView(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer


class UserCreateView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer


class UserUpdateView(generics.RetrieveUpdateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer


class UserDeleteView(generics.DestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer


# ----------------------------------------------------------------------------------------------------


class LoginView(APIView):
    def post(self, request):
        identifier = request.data.get("identifier")
        password = request.data.get("password")

        if not identifier or not password:
            return Response(
                {"error": "Toate campurile trebuie sa fie completate"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = User.objects.filter(email=identifier).first()

        if not user:
            user = User.objects.filter(username=identifier).first()

        if user:
            auth_user = authenticate(username=user.username, password=password)
            if auth_user:
                refresh = RefreshToken.for_user(auth_user)
                access_token = str(refresh.access_token)

                return Response(
                    {
                        "message": "Autentificare reusita",
                        "token": access_token,
                        "user_id": auth_user.id,
                        "first_name": auth_user.first_name,
                        "last_name": auth_user.last_name,
                        "is_superuser": auth_user.is_superuser,
                        "is_staff": auth_user.is_staff,
                    }
                )
            else:
                return Response(
                    {"error": "Parola gresita"}, status=status.HTTP_401_UNAUTHORIZED
                )
        else:
            return Response(
                {"error": "Utilizatorul nu exista"}, status=status.HTTP_404_NOT_FOUND
            )


# --------------------------------------------------------------------------------------------------


class ResetPasswordConfirmView(APIView):
    def post(self, request, *args, **kwargs):
        email = request.data.get("email")
        token = request.data.get("token")
        new_password = request.data.get("password")

        if not email or not token or not new_password:
            return Response(
                {"error": "Email, token si parola sunt obligatorii."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response(
                {"error": "Email invalid."}, status=status.HTTP_400_BAD_REQUEST
            )

        # TODO: verifica token-ul in DB asociat userului, daca e valid si neexpirat
        # Exemplu (pseudo):
        # if not PasswordResetToken.objects.filter(user=user, token=token, valid=True).exists():
        #    return Response({"error": "Token invalid sau expirat."}, status=400)

        user.set_password(new_password)
        user.save()

        # TODO: invalideaza token-ul dupa folosire

        return Response({"message": "Parola a fost resetata cu succes."})


# ----------------------------------------------------------------------------------------------------


class CheckEmailExistsView(APIView):
    def get(self, request, *args, **kwargs):
        email = request.query_params.get("email")
        if email is None:
            return Response(
                {"error": "Email is required."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({"email_exists": False})

        # Daca username == email -> trimite email pentru creare cont
        if user.username == email:
            token = str(uuid.uuid4())
            link = f"http://127.0.0.1:2002/register-confirm?token={token}"

            subject = "Creare cont - Alex Asistent Virtual"
            message = (
                f"Salut!\n\nAi solicitat crearea contului tau.\n"
                f"Click pe linkul de mai jos pentru a continua:\n\n{link}\n\n"
                f"Ignora acest mesaj daca nu ai facut tu cererea."
            )

            try:
                status_code, _ = trimite_email(
                    destinatar=email, subiect=subject, mesaj_text=message
                )

                if status_code >= 400:
                    return Response(
                        {"error": "A aparut o eroare la trimiterea emailului."},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )

                return Response(
                    {
                        "email_exists": True,
                        "message": f"Un email a fost trimis catre {email} pentru continuarea procesului.",
                    }
                )
            except Exception as e:
                return Response(
                    {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        else:
            # Daca username diferit de email → trimite email pentru resetare parola
            reset_token = str(uuid.uuid4())
            reset_link = (
                f"http://127.0.0.1:2002/reset-password-confirm?token={reset_token}"
            )

            subject = "Resetare parola - Alex Asistent Virtual"
            message = (
                f"Salut!\n\nAi solicitat resetarea parolei.\n"
                f"Click pe linkul de mai jos pentru a reseta parola:\n\n{reset_link}\n\n"
                f"Daca nu ai facut tu aceasta cerere, ignora acest mesaj."
            )

            try:
                status_code, _ = trimite_email(
                    destinatar=email, subiect=subject, mesaj_text=message
                )

                if status_code >= 400:
                    return Response(
                        {"error": "A aparut o eroare la trimiterea emailului."},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )

                return Response(
                    {
                        "email_exists": True,
                        "message": f"A fost trimis un email pentru resetarea parolei catre {email}.",
                    }
                )
            except Exception as e:
                return Response(
                    {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )


# ----------------------------------------------------------------------------------------------------


class ActivateUserView(APIView):
    def post(self, request, *args, **kwargs):
        email = request.data.get("email")
        username = request.data.get("username")
        password = request.data.get("password")

        if not email or not username or not password:
            return Response(
                {"error": "Email, username, and password are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = User.objects.get(email=email)
        except ObjectDoesNotExist:
            return Response(
                {"error": "No user with the provided email found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Daca user are username diferit de email si username este deja setat → nu permite activarea
        if user.username != email and user.username:
            return Response(
                {"error": "User already has a different username set."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user.username = username
        user.set_password(password)
        user.is_active = True
        user.save()

        return Response(
            {"success": "User activated and credentials set."},
            status=status.HTTP_200_OK,
        )


# ----------------------------------------------------------------------------------------------------


class UserProfileListView(generics.ListAPIView):
    queryset = UserProfile.objects.all()
    serializer_class = UserProfileSerializer


class UserProfileCreateView(generics.CreateAPIView):
    queryset = UserProfile.objects.all()
    serializer_class = UserProfileSerializer


class UserProfileUpdateView(generics.RetrieveUpdateAPIView):
    queryset = UserProfile.objects.all()
    serializer_class = UserProfileSerializer


class UserProfileDeleteView(generics.DestroyAPIView):
    queryset = UserProfile.objects.all()
    serializer_class = UserProfileSerializer


def get_user_profile_data(user):
    try:
        profile = UserProfile.objects.select_related(
            "facultate", "specializare", "grupa"
        ).get(user=user)
        return {
            "id_student": profile.id_student,
            "facultate": profile.facultate.nume,
            "specializare": profile.specializare.nume,
            "grupa": profile.grupa.grupa,
            "semigrupa": profile.semigrupa,
            "tip_taxa": profile.tip_taxa,
            "an_universitar": profile.an_universitar,
            "an_studiu": profile.an_studiu,
        }
    except UserProfile.DoesNotExist:
        return None


# ----------------------------------------------------------------------------------------------------


class AdevetintaListView(generics.ListAPIView):
    queryset = Adevetinta.objects.all()
    serializer_class = AdevetintaSerializer


class AdevetintaCreateView(generics.CreateAPIView):
    queryset = Adevetinta.objects.all()
    serializer_class = AdevetintaSerializer


class AdevetintaUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Adevetinta.objects.all()
    serializer_class = AdevetintaSerializer


class AdevetintaDeleteView(generics.DestroyAPIView):
    queryset = Adevetinta.objects.all()
    serializer_class = AdevetintaSerializer


# ----------------------------------------------------------------------------------------------------


class ConversationChat(APIView):
    # Endpoint public – nu este necesara autentificarea
    permission_classes = [AllowAny]

    # Instantierea clasei RAG (pentru intrebari legate de facultate)
    rag = RAG()

    def post(self, request):
        try:
            # Extrage intrebarea trimisa de utilizator
            question = request.data.get("message")

            # Elimina diacriticele din intrebare
            question = remove_diacritics(question)

            # Verifica daca intrebarea este goala
            if not question:
                return Response(
                    {"error": "Mesajul nu poate fi gol."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Comanda pentru listarea comenzilor disponibile
            if question.strip().lower() == "comenzi":
                answer_text = (
                    "Comenzile disponibile sunt:\n"
                    "- Genereaza o adeverinta cu motivul [motiv]\n"
                    "  Exemplu: Genereaza o adeverinta cu motivul angajare\n"
                )

                # Salveaza comanda in istoric daca utilizatorul este autentificat
                if request.user and request.user.is_authenticated:
                    ConversationHistory.objects.create(
                        user=request.user, question=question, answer=answer_text
                    )

                return Response({"answer": answer_text}, status=status.HTTP_200_OK)

            # Generarea unei adeverinte pentru utilizatori autentificati
            if (
                request.user
                and request.user.is_authenticated
                and question.lower().startswith("genereaza o adeverinta cu motivul")
            ):
                user = request.user
                trigger = "genereaza o adeverinta cu motivul"
                motiv = question[len(trigger) :].strip()

                # Verifica daca motivul a fost specificat
                if not motiv:
                    return Response(
                        {"answer": "Te rog specifica motivul pentru adeverinta."},
                        status=status.HTTP_200_OK,
                    )

                try:
                    # Obtine profilul utilizatorului
                    profile = UserProfile.objects.get(user=user)
                except UserProfile.DoesNotExist:
                    return Response(
                        {"error": "Profilul utilizatorului nu a fost gasit."},
                        status=status.HTTP_404_NOT_FOUND,
                    )

                # Genereaza datele pentru noua adeverinta
                numar = Adevetinta.objects.count() + 1
                data_emitere = timezone.now().date()

                # Creeaza obiectul Adeverinta in baza de date
                adeverinta = Adevetinta.objects.create(
                    user=user,
                    last_name=user.last_name,
                    first_name=user.first_name,
                    an_universitar=profile.an_universitar,
                    an_studiu=profile.an_studiu,
                    specializare=profile.specializare.nume,
                    tip_taxa=profile.tip_taxa,
                    motiv=motiv,
                    numar=numar,
                    expires_at=timezone.now() + relativedelta(years=1),
                )

                # Imparte motivul in doua linii
                motiv = adeverinta.motiv or ""
                max_len = 70
                split_pos = motiv.rfind(" ", 0, max_len)
                if split_pos == -1:
                    split_pos = max_len

                motiv1 = motiv[:split_pos]
                motiv2 = motiv[split_pos:].strip()

                # Creeaza contextul pentru template-ul HTML
                context = {
                    "numar": adeverinta.numar,
                    "data": data_emitere.strftime("%d.%m.%Y"),
                    "nume": adeverinta.last_name,
                    "prenume": adeverinta.first_name,
                    "anUniversitar": adeverinta.an_universitar,
                    "anul": adeverinta.an_studiu,
                    "specializare": adeverinta.specializare,
                    "regim": adeverinta.tip_taxa,
                    "motiv1": motiv1,
                    "motiv2": motiv2,
                }

                # Genereaza continutul HTML al adeverintei
                html_string = render_to_string("adeverinta/adeverinta.html", context)

                # Genereaza fisierul PDF pe baza HTML-ului
                pdf_bytes = HTML(
                    string=html_string, base_url=request.build_absolute_uri("/")
                ).write_pdf()

                # Creeaza un nume unic pentru fisierul PDF
                unique_id = str(uuid.uuid4())
                filename = f"adeverinta_{unique_id}.pdf"
                pdf_path = os.path.join(settings.MEDIA_ROOT, filename)

                # Creeaza folderul MEDIA daca nu exista
                if not os.path.exists(settings.MEDIA_ROOT):
                    os.makedirs(settings.MEDIA_ROOT)

                # Scrie fisierul PDF pe disc
                with open(pdf_path, "wb") as f:
                    f.write(pdf_bytes)

                # Genereaza URL-ul public pentru fisierul PDF
                pdf_url = request.build_absolute_uri(settings.MEDIA_URL + filename)

                answer_text = f"Adeverinta a fost creata si e valabila un an: {pdf_url}"

                # Salveaza in istoric intrebarea si raspunsul
                if request.user and request.user.is_authenticated:
                    ConversationHistory.objects.create(
                        user=request.user, question=question, answer=answer_text
                    )

                return Response({"answer": answer_text}, status=status.HTTP_200_OK)

            # Blocare generare adeverinta pentru utilizatori neautentificati
            if question.lower().startswith("genereaza o adeverinta cu motivul"):
                answer_text = f"Doar persoanele autentificate pot genera o adeverinta!"
                return Response({"answer": answer_text}, status=status.HTTP_200_OK)

            # Preluare datele personale pentru utilizatorul autentificat
            user_profile_data = None
            if request.user and request.user.is_authenticated:
                user_profile_data = get_user_profile_data(request.user)

            # Preluare istoric conversatie pentru utilizatorul autentificat
            conversation_history = None
            if request.user and request.user.is_authenticated:
                conversation_history = ConversationHistory.objects.filter(
                    user=request.user
                ).order_by("id")[:10]

            # Apeleaza RAG pentru a genera raspunsul pe baza intrebarii si a contextului
            answer = self.rag.get_response(
                question, user_profile_data, conversation_history
            )

            # Salveaza conversatia in istoric daca utilizatorul este autentificat
            if request.user and request.user.is_authenticated:
                logger.info(f"User {request.user.username} is authenticated.")
                ConversationHistory.objects.create(
                    user=request.user, question=question, answer=answer
                )

            # Returneaza raspunsul final
            return Response({"answer": answer}, status=status.HTTP_200_OK)

        except Exception as e:
            # Tratare erori interne de server
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ----------------------------------------------------------------------------------------------------


class ConversationHistoryListView(generics.ListAPIView):
    queryset = ConversationHistory.objects.all()
    serializer_class = ConversationHistorySerializer


class ConversationHistoryCreateView(generics.CreateAPIView):
    queryset = ConversationHistory.objects.all()
    serializer_class = ConversationHistorySerializer


class ConversationHistoryUpdateView(generics.RetrieveUpdateAPIView):
    queryset = ConversationHistory.objects.all()
    serializer_class = ConversationHistorySerializer


class ConversationHistoryDeleteView(generics.DestroyAPIView):
    queryset = ConversationHistory.objects.all()
    serializer_class = ConversationHistorySerializer


class ConversationHistoryFilteredView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        if request.user.id != pk:
            return Response(
                {
                    "detail": "Nu ai permisiunea sa accesezi mesajele acestui utilizator."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            conversations = ConversationHistory.objects.filter(user__id=pk)

            serializer = ConversationHistorySerializer(conversations, many=True)

            return Response(serializer.data)

        except ConversationHistory.DoesNotExist:
            return Response(
                {"detail": "Utilizatorul nu are mesaje."},
                status=status.HTTP_404_NOT_FOUND,
            )


# ----------------------------------------------------------------------------------------------------


class DownloadFilesView(APIView):
    def post(self, request):
        url = "https://fiesc.usv.ro/docs/"

        try:
            logger.info(
                "Incepe descarcarea fisierelor .pdf/.doc/.docx de pe site: %s", url
            )
            response = requests.get(url)
            response.raise_for_status()

            soup = BeautifulSoup(response.text, "html.parser")

            links = soup.find_all("a", href=True)
            file_urls = [
                link["href"]
                for link in links
                if link["href"].endswith((".pdf", ".doc", ".docx"))
            ]

            if not file_urls:
                logger.warning("Nu s-au gasit fisiere .pdf/.doc/.docx pe site.")
                return Response(
                    {
                        "status": "warning",
                        "message": "Nu s-au gasit fisiere .pdf/.doc/.docx pe site.",
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            base_dir = settings.BASE_DIR
            pdf_dir = os.path.join(base_dir, "pdfs")
            docs_dir = os.path.join(base_dir, "docs")

            if os.path.exists(pdf_dir):
                shutil.rmtree(pdf_dir)
                logger.info("Directorul 'pdfs' a fost sters.")
            if os.path.exists(docs_dir):
                shutil.rmtree(docs_dir)
                logger.info("Directorul 'docs' a fost sters.")

            os.makedirs(pdf_dir, exist_ok=True)
            os.makedirs(docs_dir, exist_ok=True)

            for file_url in file_urls:
                try:
                    if not file_url.startswith("http"):
                        file_url = "https://fiesc.usv.ro" + file_url
                    file_name = file_url.split("/")[-1]
                    ext = os.path.splitext(file_name)[1].lower()

                    save_dir = pdf_dir if ext == ".pdf" else docs_dir
                    file_path = os.path.join(save_dir, file_name)

                    logger.info(f"Se descarca fisierul: {file_name}")
                    with requests.get(file_url, stream=True) as file_response:
                        file_response.raise_for_status()
                        with open(file_path, "wb") as f:
                            for chunk in file_response.iter_content(chunk_size=8192):
                                f.write(chunk)

                    logger.info(f"Fisierul {file_name} a fost descarcat cu succes.")
                except requests.exceptions.RequestException as e:
                    logger.error(f"Eroare la descarcarea fisierului {file_name}: {e}")
                    continue

            return Response(
                {
                    "status": "success",
                    "message": "Fisierele au fost descarcate cu succes.",
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            logger.error(f"erroare {e}")
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ----------------------------------------------------------------------------------------------------


@method_decorator(csrf_exempt, name="dispatch")
class ImportUsersFromExcelView(APIView):

    def post(self, request, *args, **kwargs):
        if "file" not in request.FILES:
            return JsonResponse(
                {"error": "Metoda trebuie POST cu fisierul Excel"}, status=400
            )

        excel_file = request.FILES["file"]
        wb = load_workbook(filename=excel_file, read_only=True)
        ws = wb.active

        an_universitar = ws["A5"].value

        row = 9
        while True:
            id_student = ws[f"B{row}"].value
            if id_student is None:
                break

            nume = ws[f"C{row}"].value
            prenume = ws[f"D{row}"].value
            email = ws[f"E{row}"].value
            semitrupa_code = ws[f"F{row}"].value
            tip_taxa = ws[f"G{row}"].value

            if semitrupa_code and len(semitrupa_code) >= 5:
                try:
                    id_facultate = int(semitrupa_code[0])
                    id_specializare = int(semitrupa_code[1])
                    an_studiu_gr = int(semitrupa_code[2])
                    grupa_nr = int(semitrupa_code[3])
                    semigrupa_char = semitrupa_code[4]
                except ValueError:
                    return JsonResponse(
                        {"error": f"Semitrupa code invalid la randul {row}"}, status=400
                    )
            else:
                return JsonResponse(
                    {"error": f"Semitrupa code invalid la randul {row}"}, status=400
                )

            try:
                facultate = Facultate.objects.get(id=id_facultate)
                specializare = Specializare.objects.get(
                    id=id_specializare, facultate=facultate
                )
                grupa = Grupa.objects.get(
                    facultate=facultate,
                    specializare=specializare,
                    an_studiu=an_studiu_gr,
                    grupa=grupa_nr,
                    semigrupa=semigrupa_char,
                )
            except Facultate.DoesNotExist:
                return JsonResponse(
                    {
                        "error": f"Facultate cu id {id_facultate} nu exista la randul {row}"
                    },
                    status=400,
                )
            except Specializare.DoesNotExist:
                return JsonResponse(
                    {
                        "error": f"Specializare cu id {id_specializare} nu exista la randul {row}"
                    },
                    status=400,
                )
            except Grupa.DoesNotExist:
                return JsonResponse(
                    {"error": f"Grupa nu exista la randul {row}"}, status=400
                )

            user, created = User.objects.get_or_create(
                username=email,
                defaults={"first_name": prenume, "last_name": nume, "email": email},
            )

            if not created:
                user.first_name = prenume
                user.last_name = nume
                user.email = email
                user.save()

            user_profile, _ = UserProfile.objects.get_or_create(user=user)
            user_profile.id_student = id_student
            user_profile.facultate = facultate
            user_profile.specializare = specializare
            user_profile.grupa = grupa
            user_profile.semigrupa = semigrupa_char
            user_profile.tip_taxa = tip_taxa
            user_profile.an_universitar = an_universitar
            user_profile.an_studiu = an_studiu_gr
            user_profile.save()

            row += 1

        return JsonResponse(
            {"status": "Succes", "message": f"Importate {row-9} utilizatori"}
        )


# --------------------------------------------------------------------------------------------------

load_dotenv()

SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY")
FROM_EMAIL = os.getenv("SENDGRID_FROM_EMAIL")


def trimite_email(destinatar, subiect, mesaj_text):
    """
    Trimite un email simplu folosind SendGrid Web API.

    :param destinatar: Emailul destinatarului
    :param subiect: Subiectul emailului
    :param mesaj_text: Textul emailului (de exemplu, cu linkul de resetare)
    :return: status_code și răspunsul text de la SendGrid
    """
    if not SENDGRID_API_KEY or not FROM_EMAIL:
        raise ValueError("SENDGRID_API_KEY sau FROM_EMAIL nu este setat în .env")

    url = "https://api.sendgrid.com/v3/mail/send"
    headers = {
        "Authorization": f"Bearer {SENDGRID_API_KEY}",
        "Content-Type": "application/json",
    }

    payload = {
        "personalizations": [{"to": [{"email": destinatar}], "subject": subiect}],
        "from": {"email": FROM_EMAIL},
        "content": [{"type": "text/plain", "value": mesaj_text}],
    }

    response = requests.post(url, headers=headers, json=payload)
    return response.status_code, response.text


# ----------------------------------------------------------------------------------------------------


class HomeView(TemplateView):
    template_name = "home.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        facultati = Facultate.objects.all()
        specializari = Specializare.objects.all()
        grupe = Grupa.objects.all()
        users = User.objects.all()
        user_profiles = UserProfile.objects.all()
        adeverinte = Adevetinta.objects.all()
        conversation_histories = ConversationHistory.objects.all()

        context["endpoints"] = [
            ("List Facultate", reverse_lazy("facultate-list")),
            ("Create Facultate", reverse_lazy("facultate-create")),
            *[
                (
                    "Update Facultate {}".format(facultate.nume),
                    reverse_lazy("facultate-update", kwargs={"pk": facultate.pk}),
                )
                for facultate in facultati
            ],
            *[
                (
                    "Delete Facultate {}".format(facultate.nume),
                    reverse_lazy("facultate-delete", kwargs={"pk": facultate.pk}),
                )
                for facultate in facultati
            ],
            ("Filter Facultate", reverse_lazy("facultate-filter")),
            ("List Specializare", reverse_lazy("specializare-list")),
            ("Create Specializare", reverse_lazy("specializare-create")),
            *[
                (
                    "Update Specializare {}".format(specializare.nume),
                    reverse_lazy("specializare-update", kwargs={"pk": specializare.pk}),
                )
                for specializare in specializari
            ],
            *[
                (
                    "Delete Specializare {}".format(specializare.nume),
                    reverse_lazy("specializare-delete", kwargs={"pk": specializare.pk}),
                )
                for specializare in specializari
            ],
            ("Filter Specializare", reverse_lazy("specializare-filter")),
            ("List Grupa", reverse_lazy("grupa-list")),
            ("Create Grupa", reverse_lazy("grupa-create")),
            *[
                (
                    "Update Grupa {}".format(grupa.nume),
                    reverse_lazy("grupa-update", kwargs={"pk": grupa.pk}),
                )
                for grupa in grupe
            ],
            *[
                (
                    "Delete Grupa {}".format(grupa.nume),
                    reverse_lazy("grupa-delete", kwargs={"pk": grupa.pk}),
                )
                for grupa in grupe
            ],
            ("Filter Grupa", reverse_lazy("grupa-filter")),
            ("Login", reverse_lazy("login")),
            ("User Reset Password", reverse_lazy("user-reset-password")),
            ("User Email", reverse_lazy("user-email")),
            ("User Active", reverse_lazy("user-active")),
            ("List User", reverse_lazy("user-list")),
            ("Create User", reverse_lazy("user-create")),
            *[
                (
                    "Update User {}".format(user.username),
                    reverse_lazy("user-update", kwargs={"pk": user.pk}),
                )
                for user in users
            ],
            *[
                (
                    "Delete User {}".format(user.username),
                    reverse_lazy("user-delete", kwargs={"pk": user.pk}),
                )
                for user in users
            ],
            ("List User Profile", reverse_lazy("user-profile-list")),
            ("Create User Profile", reverse_lazy("user-profile-create")),
            *[
                (
                    "Update User Profile {}".format(user_profile.nume),
                    reverse_lazy("user-profile-update", kwargs={"pk": user_profile.pk}),
                )
                for user_profile in user_profiles
            ],
            *[
                (
                    "Delete User Profile {}".format(user_profile.nume),
                    reverse_lazy("user-profile-delete", kwargs={"pk": user_profile.pk}),
                )
                for user_profile in user_profiles
            ],
            ("List Adevetinta", reverse_lazy("adevetinta-list")),
            ("Create Adevetinta", reverse_lazy("adevetinta-create")),
            *[
                (
                    "Update Adevetinta {}".format(adevetinta.nume),
                    reverse_lazy("adevetinta-update", kwargs={"pk": adevetinta.pk}),
                )
                for adevetinta in adeverinte
            ],
            *[
                (
                    "Delete Adevetinta {}".format(adevetinta.nume),
                    reverse_lazy("adevetinta-delete", kwargs={"pk": adevetinta.pk}),
                )
                for adevetinta in adeverinte
            ],
            ("Conversation Chat", reverse_lazy("conversation-chat")),
            ("List Conversation History", reverse_lazy("conversation-history-list")),
            (
                "Create Conversation History",
                reverse_lazy("conversation-history-create"),
            ),
            *[
                (
                    "Update Conversation History {}".format(conversation_history.nume),
                    reverse_lazy(
                        "conversation-history-update",
                        kwargs={"pk": conversation_history.pk},
                    ),
                )
                for conversation_history in conversation_histories
            ],
            *[
                (
                    "Delete Conversation History {}".format(conversation_history.nume),
                    reverse_lazy(
                        "conversation-history-delete",
                        kwargs={"pk": conversation_history.pk},
                    ),
                )
                for conversation_history in conversation_histories
            ],
            *[
                (
                    "Filter Conversation History {}".format(conversation_history.nume),
                    reverse_lazy(
                        "conversation-history-filter",
                        kwargs={"pk": conversation_history.pk},
                    ),
                )
                for conversation_history in conversation_histories
            ],
            ("Download Files", reverse_lazy("download-files")),
            ("User Import", reverse_lazy("user-import")),
        ]
        return context


# ----------------------------------------------------------------------------------------------------
